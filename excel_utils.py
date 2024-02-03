import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

def write_to_excel(csv_arrays, filename):
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    with pd.ExcelWriter(filename) as writer:
        for key in csv_arrays.keys():

            # The sheet name is the filename without the extension
            sheetname = key.split('wakatime_')[1].split('_')[0]
            data = csv_arrays[key]

            # Write each dataframe to a different worksheet.
            data.to_excel(writer, sheet_name=sheetname, index=False)
    

    # Load the workbook and iterate through each cell in the sheet to format the time
    wb = load_workbook(filename)
    for key in csv_arrays.keys():
        sheetname = key.split('wakatime_')[1].split('_')[0]
        ws = wb[sheetname]
        
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                if ':' in str(cell.value):
                    # By default, Excel ranges from 0 to 23 hours, so we need to convert the time manually to show it as a time
                    cell.number_format = '[HH]:MM:SS'
                    hours, minutes = map(int, str(cell.value).split(':'))
                    total_seconds = hours * 3600 + minutes * 60
                    
                    cell.value = total_seconds / 86400
        
        # Clear the first cell of the sheet (default is Unamed: 0)
        ws['A1'].value = None

    wb.save(filename)
